# Librerias
import calendar
import os
from io import BytesIO
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ruta absoluta a la plantilla maestra .xlsm que contiene las macros VBA binarizadas
_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),       # backend/exports/
    "..",                            # backend/
    "..",                            # raiz del proyecto
    ".antigravity", "context",
    "Plantilla.xlsm"
)

# Layout del template V2 (debe coincidir exactamente con el VBA de XlsmAsistencias)
_HEADER_ROW = 4       # Fila donde van los encabezados de día (Lun 01, Mar 02, ...)
_DATA_ROW_START = 5   # Primera fila de datos de alumnos
_DATA_ROW_END = 104   # Límite máximo de alumnos (igual al rango C5:AG105 del VBA)
_COL_NICKNAME = 1     # Columna A: Apodo (visible, leído por EjecutarAsistenciaMasiva)
_COL_FULLNAME = 2     # Columna B: Nombre completo (visible)
_COL_DAYS_START = 3   # Columna C: primer día del mes
_COL_DAYS_END = 33    # Columna AG: último día posible (31 días)
_COL_ENROLLMENT_ID = 35  # Columna AI: enrollment_id (oculto, identificador unívoco)

_DIAS_SEMANA = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]
_MESES_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def generate_attendance_template(
    year: int,
    month: int,
    students_data: List[Dict[str, Any]]
) -> BytesIO:
    """
    Genera un archivo .xlsm con macros VBA incrustadas como plantilla de asistencia mensual.
    Usa Plantilla.xlsm como base (keep_vba=True) para preservar el vbaProject.bin.

    students_data debe contener diccionarios con:
      - enrollment_id (int): identificador unívoco de la inscripción
      - nickname       (str): apodo corto del alumno (leído por la macro Express)
      - full_name      (str): nombre completo del alumno
    """
    template_path = os.path.normpath(_TEMPLATE_PATH)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró la plantilla maestra en: {template_path}"
        )

    # Abrir la plantilla preservando las macros VBA binarias
    wb = openpyxl.load_workbook(template_path, keep_vba=True)

    # 1. Seleccionar y renombrar la hoja de asistencias (siempre la primera)
    ws = wb.worksheets[0]
    sheet_name = f"Asistencias {year}-{month:02d}"
    ws.title = sheet_name

    # Actualizar la referencia al nombre de la hoja en la hoja "Express" (Hoja 2)
    # La macro XlsmExpress.vba referencia el nombre exacto de la hoja.
    # openpyxl no puede editar código VBA directamente, pero el nombre de hoja
    # queda ligado via el índice del worksheet (ThisWorkbook.Worksheets(1)),
    # por lo que el rename no rompe las macros.

    # 2. Limpiar datos previos de la plantilla (filas 4 a 104, cols A-AI)
    for row in range(_HEADER_ROW, _DATA_ROW_END + 1):
        for col in range(1, _COL_ENROLLMENT_ID + 1):
            ws.cell(row=row, column=col).value = None

    # 3. Estilos de encabezado (consistentes con el diseño original de la plantilla)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # 4. Escribir encabezados de días en Fila 4 (cols C = 3 a AG = 33)
    _, num_days = calendar.monthrange(year, month)

    # Primero limpiar toda la fila de encabezados de días
    for col in range(_COL_DAYS_START, _COL_DAYS_END + 1):
        ws.cell(row=_HEADER_ROW, column=col).value = None

    # Escribir solo los días hábiles del mes
    dias_del_mes = []  # Lista de (day_num, col_index)
    col_cursor = _COL_DAYS_START
    for day in range(1, num_days + 1):
        weekday = calendar.weekday(year, month, day)
        if weekday < 5:  # Lunes (0) a Viernes (4) únicamente
            cell = ws.cell(row=_HEADER_ROW, column=col_cursor)
            cell.value = f"{_DIAS_SEMANA[weekday]} {day:02d}"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin
            dias_del_mes.append((day, col_cursor))
            col_cursor += 1

    # 5. Encabezados fijos de columnas A y B en Fila 4
    for col, label in [(_COL_NICKNAME, "APODO"), (_COL_FULLNAME, "NOMBRE COMPLETO")]:
        cell = ws.cell(row=_HEADER_ROW, column=col)
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # 6. Llenar filas de alumnos a partir de Fila 5
    alt_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    data_font = Font(color="E5E7EB", size=9)
    left_align = Alignment(horizontal="left", vertical="center")

    for i, student in enumerate(students_data):
        row = _DATA_ROW_START + i
        if row > _DATA_ROW_END:
            break  # Límite de 100 alumnos por hoja

        fill = alt_fill if i % 2 == 0 else PatternFill(
            start_color="1A2332", end_color="1A2332", fill_type="solid"
        )

        # Columna A: Apodo (leído por la macro de Asistencia Express para identificar filas)
        c_nick = ws.cell(row=row, column=_COL_NICKNAME)
        c_nick.value = student["nickname"]
        c_nick.fill = fill
        c_nick.font = data_font
        c_nick.alignment = center
        c_nick.border = thin

        # Columna B: Nombre completo
        c_full = ws.cell(row=row, column=_COL_FULLNAME)
        c_full.value = student["full_name"]
        c_full.fill = fill
        c_full.font = data_font
        c_full.alignment = left_align
        c_full.border = thin

        # Columnas C–AG: Valor inicial 0 (invisible) — la macro lo reemplaza con Time al escribirlo
        for _, col_idx in dias_del_mes:
            c_day = ws.cell(row=row, column=col_idx)
            c_day.value = 0
            c_day.number_format = ';;'  # Formato que oculta el 0 visualmente
            c_day.fill = fill
            c_day.font = data_font
            c_day.alignment = center
            c_day.border = thin

        # Columna AI (35): enrollment_id oculto — identificador unívoco para el parser del backend
        c_enr = ws.cell(row=row, column=_COL_ENROLLMENT_ID)
        c_enr.value = student["enrollment_id"]

    # 7. Ocultar columna AI (enrollment_id) y ajustar anchos
    ws.column_dimensions[get_column_letter(_COL_ENROLLMENT_ID)].hidden = True
    ws.column_dimensions[get_column_letter(_COL_NICKNAME)].width = 14
    ws.column_dimensions[get_column_letter(_COL_FULLNAME)].width = 34
    for _, col_idx in dias_del_mes:
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    # 8. Congelar paneles: la primera columna de días queda fija al scrollear horizontalmente
    first_day_letter = get_column_letter(_COL_DAYS_START)
    ws.freeze_panes = f"{first_day_letter}{_DATA_ROW_START}"

    # 9. Serializar como .xlsm (las macros quedan preservadas gracias a keep_vba=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
