# Librerias
from datetime import date, time
from io import BytesIO
from sqlalchemy.orm import Session
import openpyxl
# Importar directorios del proyecto
from repositories import (
    attendance_search_by_enrollment_and_date as search_by_enrollment_and_date,
    attendance_create as create,
    attendance_update as update)

# ─── Constantes del layout V2 (deben coincidir con excel_template_generator.py) ──
_HEADER_ROW       = 4   # Fila con encabezados de día ("Lun 01", "Mar 02", ...)
_DATA_ROW_START   = 5   # Primera fila de datos de alumnos
_COL_NICKNAME     = 1   # Columna A: Apodo (no se usa para el import, solo referencia)
_COL_FULLNAME     = 2   # Columna B: Nombre completo (no se usa para el import)
_COL_DAYS_START   = 3   # Columna C: primer día del mes
_COL_DAYS_END     = 33  # Columna AG: último día posible
_COL_ENROLLMENT   = 35  # Columna AI: enrollment_id (identificador unívoco, oculto)


def parse_arrival_time(cell_value) -> time | None:
    """
    Convierte el valor de una celda del Excel a un objeto time.
    Acepta:
      - Objetos time/datetime de openpyxl (celda con formato hora)
      - Strings: "HH:MM", "HH:MM:SS", "H:MM AM/PM"
      - 0 / 0.0 / None / cadena vacía → devuelve None (alumno AUSENTE)
    En el template V2, las celdas de días se inicializan con 0 (valor invisible).
    La macro VBA solo reemplaza ese 0 con Time() cuando el profe escribe 1.
    """
    if cell_value is None:
        return None

    # Valor 0 o 0.0 → la macro no lo tocó → alumno ausente
    if isinstance(cell_value, (int, float)) and cell_value == 0:
        return None

    if isinstance(cell_value, time):
        return cell_value

    # openpyxl a veces devuelve datetime para celdas con formato hora
    from datetime import datetime as dt
    if isinstance(cell_value, dt):
        return cell_value.time()

    # Fracción decimal de Excel (ej: 0.3229 = 07:45:00)
    # openpyxl ya los convierte a time cuando data_only=True, pero por seguridad:
    if isinstance(cell_value, float) and 0.0 < cell_value < 1.0:
        total_seconds = int(round(cell_value * 86400))
        h, rem = divmod(total_seconds, 3600)
        m, s = divmod(rem, 60)
        return time(h % 24, m, s)

    raw = str(cell_value).strip()
    if not raw or raw == "0":
        return None

    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            from datetime import datetime as dt
            return dt.strptime(raw, fmt).time()
        except ValueError:
            continue

    return None  # Valor no reconocido → se trata como ausente


def _extract_year_month_from_sheet(ws) -> tuple[int, int]:
    """
    Extrae año y mes desde el nombre de la hoja.
    Formato esperado del título: "Asistencias YYYY-MM"
    """
    title = ws.title  # Ej: "Asistencias 2026-04"
    parts = title.split()  # ["Asistencias", "2026-04"]
    if len(parts) >= 2:
        date_part = parts[-1]  # "2026-04"
        if "-" in date_part:
            year_str, month_str = date_part.split("-", 1)
            try:
                return int(year_str), int(month_str)
            except ValueError:
                pass
    raise ValueError(
        f"No se pudo extraer año/mes del nombre de la hoja: '{title}'. "
        "El formato esperado es 'Asistencias YYYY-MM'."
    )


def _extract_day_columns(ws) -> dict[int, int]:
    """
    Lee la fila de encabezados de días (HEADER_ROW = 4), columnas C–AG.
    Devuelve un dict {col_index: numero_de_dia} solo para columnas con encabezado válido.
    Formato del encabezado: "Lun 01", "Mar 02", "Mie 15", etc.
    """
    import re
    day_columns: dict[int, int] = {}
    for c in range(_COL_DAYS_START, _COL_DAYS_END + 1):
        val = ws.cell(row=_HEADER_ROW, column=c).value
        if val is None:
            continue
        match = re.search(r'(\d+)', str(val).strip())
        if match:
            day_columns[c] = int(match.group(1))
    return day_columns


def process_attendance_excel(file_content: bytes, db: Session) -> dict:
    """
    Parsea un archivo .xlsm (Template V2) y realiza UPSERT de asistencias.

    Layout V2:
      - Nombre de hoja: "Asistencias YYYY-MM"  (fuente del mes/año)
      - Fila 4: encabezados de día en columnas C (3) a AG (33)
      - Fila 5+: datos de alumnos
          Col A (1): Apodo
          Col B (2): Nombre completo
          Col C–AG:  Hora de llegada (time) o 0 (=ausente, la macro no lo tocó)
          Col AI (35): enrollment_id  (identificador unívoco, oculto)
    """
    # keep_vba=True para que openpyxl no rechace el formato .xlsm
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True, keep_vba=True)
    ws = wb.active

    # ─── 1. Extraer año y mes desde el nombre de la hoja ─────────────────────
    year, month = _extract_year_month_from_sheet(ws)

    # ─── 2. Identificar columnas de días en Fila 4 ────────────────────────────
    day_columns = _extract_day_columns(ws)
    if not day_columns:
        raise ValueError(
            "No se encontraron columnas de días en la fila 4 (columnas C–AG). "
            "Confirme que la plantilla tiene el formato V2 correcto."
        )

    total_insertados   = 0
    total_actualizados = 0
    total_omitidos     = 0

    # ─── 3. Procesar filas de alumnos (desde Fila 5) ─────────────────────────
    for r in range(_DATA_ROW_START, ws.max_row + 1):
        # El enrollment_id está en la columna AI (35), oculta
        enrollment_id_raw = ws.cell(row=r, column=_COL_ENROLLMENT).value

        # Si la columna AI está vacía nos detuvimos en las filas de datos reales
        if not enrollment_id_raw:
            # Doble-check: si el apodo también está vacío es fin de tabla
            if not ws.cell(row=r, column=_COL_NICKNAME).value:
                break
            # Fila sin enrollment_id pero con apodo → omitir y continuar
            total_omitidos += 1
            continue

        try:
            enrollment_id = int(enrollment_id_raw)
        except (ValueError, TypeError):
            total_omitidos += 1
            continue

        for col_index, day in day_columns.items():
            cell = ws.cell(row=r, column=col_index)
            arrival = parse_arrival_time(cell.value)

            # Extraer nota desde comentario de celda si existe
            notes: str | None = cell.comment.text if cell.comment else None

            # Si ni hay hora ni nota, la celda está «ausente sin registro» → omitir
            if arrival is None and notes is None:
                continue

            attendance_date = date(year, month, day)
            existing = search_by_enrollment_and_date(db, enrollment_id, attendance_date)

            if existing:
                changed = (existing.arrival_time != arrival) or (existing.notes != notes)
                if changed:
                    update(db, existing.id,
                           attendance_date=attendance_date,
                           arrival_time=arrival,
                           notes=notes)
                    total_actualizados += 1
                else:
                    total_omitidos += 1
            else:
                create(db,
                       attendance_date=attendance_date,
                       arrival_time=arrival,
                       notes=notes,
                       enrollment_id=enrollment_id)
                total_insertados += 1

    return {
        "success": True,
        "inserted": total_insertados,
        "updated":  total_actualizados,
        "skipped":  total_omitidos,
        "received_month": month,
        "received_year":  year,
        "logs": []
    }
